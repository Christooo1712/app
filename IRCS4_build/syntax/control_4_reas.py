import pandas as pd
import glob
import os
from concurrent.futures import ProcessPoolExecutor
import re
from openpyxl import load_workbook

columns_to_sum_argo = [
    'prm_inc','lrc_cl_ins','lrc_cl_inv','r_exp_m','r_acq_cost','cov_units','DAC_COV_UNITS','dac','exp_acq',
    'lrc_cl_ins_dth','lrc_cl_inv_dth','lrc_cl_inv_surr','lrc_cl_inv_mat','lrc_cl_inv_ann'
]
columns_to_sum_rafm = ['prm_inc','lrc_cl_ins','cov_units','pv_reins_clm','lrc_cl_ins_dth']
cols_to_compare = ['prm_inc','lrc_cl_ins','cov_units','dac_cov_units','lrc_cl_ins_dth']
target_sheets = ['extraction IDR', 'extraction USD']
global_filter_rafm = None

def parse_numeric_fast(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val)

    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ['none', 'nan', 'n/a', '-', '--']:
            return None
        s = s.replace('\xa0', '').replace(' ', '').replace('\u202f','')
        s = s.replace('−', '-')
        s = re.sub(r'[^\d,.\-()%]', '', s)
        is_percent = s.endswith('%')
        if is_percent:
            s = s[:-1]
        is_negative = False
        if s.startswith('(') and s.endswith(')'):
            is_negative = True
            s = s[1:-1]

        comma_count = s.count(',')
        dot_count = s.count('.')

        try:
            if comma_count > 1 and dot_count == 1 and s.rfind('.') > s.rfind(','):
                result = float(s.replace(',', ''))
            elif comma_count == 1 and dot_count > 0 and s.rfind(',') > s.rfind('.'):
                result = float(s.replace('.', '').replace(',', '.'))
            elif dot_count == 0 and comma_count == 1:
                result = float(s.replace(',', '.'))
            elif dot_count > 1 and comma_count == 0:
                result = float(s.replace('.', ''))
            elif comma_count == 0 and dot_count <= 1:
                result = float(s)
            else:
                result = float(s.replace(',', '').replace('.', ''))

            if is_negative:
                result = -result
            if is_percent:
                result /= 100.0

            return result

        except ValueError:
            return None

    try:
        return float(val)
    except:
        return None


def process_argo_file(file_path):
    file_name_argo = os.path.splitext(os.path.basename(file_path))[0]
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        sheet = wb['Sheet1']
        
        data = list(sheet.values)
        if not data:
            wb.close()
            print(f"❌ File {file_name_argo} kosong")
            return {'File_Name': file_name_argo}
        
        header = [str(h).strip().lower() for h in data[0]]
        col_index = {}
        for col in columns_to_sum_argo:
            try:
                idx = header.index(col.lower())
                col_index[col] = idx
            except ValueError:
                print(f"⚠️ Kolom '{col}' tidak ditemukan di file {file_name_argo}")
        
        sums = {col: 0 for col in col_index}
        row_count = 0
        parsed_count = {col: 0 for col in col_index}
        skipped_count = {col: 0 for col in col_index}
    
        for row_idx, row in enumerate(data[1:], start=2):
            row_count += 1
            for col, idx in col_index.items():
                if idx < len(row):
                    val = row[idx]
                    parsed_val = parse_numeric_fast(val)
                    if parsed_val is not None:
                        sums[col] += parsed_val
                        parsed_count[col] += 1
                    else:
                        skipped_count[col] += 1
        
        wb.close()
    except Exception as e:
        print(f"❌ Gagal proses {file_name_argo}: {e}")
        sums = {}
    
    sums['File_Name'] = file_name_argo
    return sums

def process_rafm_file(entry):
    file_path, file_name = entry
    total_sums = {col: 0 for col in columns_to_sum_rafm}

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
    except Exception as e:
        print(f"❌ Tidak bisa membuka file {file_name}: {e}")
        return {**total_sums, 'File_Name': file_name}

    for sheet_name in target_sheets:
        try:
            if sheet_name not in wb.sheetnames:
                continue

            sheet = wb[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header = None
            for _ in range(20):
                raw = next(rows, [])
                cleaned = [str(h).strip().lower() if h else '' for h in raw]
                if 'goc' in cleaned:
                    header = cleaned
                    break

            if not header:
                print(f"⚠️ Kolom 'GOC' tidak ditemukan dalam 20 baris pertama di sheet {sheet_name} file {file_name}, dilewati.")
                continue
            data_start = []
            for _ in range(3):
                peek = next(rows, [])
                if any(peek):
                    data_start = [peek]
                    break
            col_index = {}
            lower_targets = [c.lower() for c in columns_to_sum_rafm]
            for i, col in enumerate(header):
                if col in lower_targets:
                    col_index[col] = i
            for row in data_start + list(rows):
                for col in columns_to_sum_rafm:
                    idx = col_index.get(col.lower())
                    if idx is not None and idx < len(row):
                        parsed_val = parse_numeric_fast(row[idx])
                        if parsed_val is not None and parsed_val != 0:
                            total_sums[col] += parsed_val

        except Exception as e:
            print(f"   ❌ Error processing sheet {sheet_name} file {file_name}: {e}")
            continue

    wb.close()
    total_sums['File_Name'] = file_name
    return total_sums

def main(params):
    global columns_to_sum_argo, columns_to_sum_rafm, cols_to_compare, target_sheets

    input_excel = params['input excel']

    excel_file = pd.ExcelFile(input_excel)
    code = pd.read_excel(excel_file, sheet_name='Code')
    sign_logic = pd.read_excel(excel_file, sheet_name='Sign Logic')
    control = pd.read_excel(excel_file, sheet_name='Control')
    file_path_df = pd.read_excel(excel_file, sheet_name='File Path')
    excel_file.close()

    path_map = dict(zip(file_path_df['Name'].str.lower(), file_path_df['File Path']))

    folder_path_argo = path_map.get('argo', '')
    folder_path_rafm = path_map.get('rafm', '')

    argo_files_from_code = set(code['ARGO File Name'].astype(str).str.strip().str.lower())
    rafm_files_from_code = set(code['RAFM File Name'].astype(str).str.strip().str.lower())

    file_paths_argo = [
        f for f in glob.glob(os.path.join(folder_path_argo, '*.xlsx'))
        if os.path.splitext(os.path.basename(f))[0].lower() in argo_files_from_code
        and not os.path.basename(f).startswith('~$')
    ]

    file_paths_rafm = [
        f for f in glob.glob(os.path.join(folder_path_rafm, '*.xlsx'))
        if os.path.splitext(os.path.basename(f))[0].lower() in rafm_files_from_code
        and not os.path.basename(f).startswith('~$')
    ]
    
    optimal_workers = min(os.cpu_count() or 4, max(len(file_paths_argo), 1))
    
    with ProcessPoolExecutor(max_workers=optimal_workers) as executor:
        summary_rows_argo = list(filter(None, executor.map(process_argo_file, file_paths_argo)))

    cf_argo = pd.DataFrame(summary_rows_argo)
    cf_argo = cf_argo[['File_Name'] + [col for col in cf_argo.columns if col != 'File_Name']]
    cf_argo = cf_argo.rename(columns={'File_Name': 'ARGO File Name', 'DAC_COV_UNITS': 'dac_cov_units'})
    
    mask = code['RAFM File Name'].astype(str).str.contains('_ori', regex=True, na=False)
    code = code[~mask].copy()
    
    cf_argo = pd.merge(code, cf_argo, on='ARGO File Name', how='left')
    
    columns_to_drop = [col for col in ['RAFM File Name', 'UVSG File Name'] if col in cf_argo.columns]
    if columns_to_drop:
        cf_argo = cf_argo.drop(columns=columns_to_drop)

    file_entries = [(f, os.path.splitext(os.path.basename(f))[0]) for f in file_paths_rafm]

    with ProcessPoolExecutor(max_workers=optimal_workers) as executor:
        results = list(executor.map(process_rafm_file, file_entries))

    summary_rows_rafm = [result for result in results if result]

    cf_rafm_1 = pd.DataFrame(summary_rows_rafm)

    if not cf_rafm_1.empty and 'File_Name' in cf_rafm_1.columns:
        cols = ['File_Name'] + [col for col in cf_rafm_1.columns if col != 'File_Name']
        cf_rafm_1 = cf_rafm_1[cols]
    cf_rafm_1 = cf_rafm_1.rename(columns={'File_Name': 'RAFM File Name'})
    cf_rafm_1 = cf_rafm_1.groupby('RAFM File Name', as_index=False).first()
    cf_rafm_merge = pd.merge(code, cf_rafm_1, on="RAFM File Name", how="left").fillna(0)

    numeric_cols = cf_rafm_merge.select_dtypes(include='number').columns
    sum_rows = cf_rafm_merge[cf_rafm_merge['RAFM File Name'].str.contains("SUM_", na=False)]

    for idx, row in sum_rows.iterrows():
        rafm_value = row['RAFM File Name']

        if 'SUM_' in rafm_value:
            keyword = rafm_value.split('SUM_')[-1]
            pattern_search = re.escape(keyword).replace("-", "[-_]?")
            matched_rows = cf_rafm_merge[cf_rafm_merge['ARGO File Name'].str.contains(
                pattern_search, case=False, regex=True, na=False)]

            total_values = matched_rows[numeric_cols].sum()

            for col in numeric_cols:
                cf_rafm_merge.at[idx, col] = total_values[col]

    columns_to_drop = [col for col in ['ARGO File Name', 'UVSG File Name'] if col in cf_rafm_merge.columns]
    if columns_to_drop:
        cf_rafm = cf_rafm_merge.drop(columns=columns_to_drop)
    else:
        cf_rafm = cf_rafm_merge.copy()

    cf_rafm['dac_cov_units'] = cf_rafm['cov_units']
    
    final = code.copy()
    for col in cols_to_compare:
        if col not in code.columns:
            final[col] = pd.NA

    logic_row = sign_logic.iloc[0]
    valid_cols = [col for col in logic_row.index if col in cf_argo.columns]
    
    def check_sign(val, logic_sign):
        if pd.isna(val):
            return 0
        if logic_sign == 1:
            return 1 if val < 0 else 0
        elif logic_sign == "-":
            return 0  
        elif logic_sign == -1:
            return 1 if val > 0 else 0 
        return 0 

    check_sign_summary_row = {
        col: cf_argo[col].apply(lambda val: check_sign(val, logic_row[col])).sum()
        for col in valid_cols
    }

    for col in cf_argo.columns:
        if col not in check_sign_summary_row:
            check_sign_summary_row[col] = None
    
    check_sign_summary = pd.DataFrame([check_sign_summary_row])[cf_argo.columns]
    cf_argo = pd.concat([cf_argo, check_sign_summary], ignore_index=True)
    check_sign_total = sum(val for val in check_sign_summary_row.values() if isinstance(val, (int, float)))
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = check_sign_total
    
    index_labels = list(range(1, len(cf_argo))) + ['check sign']
    cf_argo.insert(0, 'No', index_labels)
    cf_argo = pd.concat([cf_argo, sign_logic], ignore_index=True)
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = 'Sign Logic'
    
    index_labels_rafm = list(range(1, len(cf_rafm)+1))
    cf_rafm.insert(0, 'No', index_labels_rafm)
    
    mask = final['RAFM File Name'].astype(str).str.contains('_ori', regex=True, na=False)
    final = final[~mask].copy()
    
    index_labels_final = list(range(1, len(final)+1))
    final.insert(0, 'No', index_labels_final)

    control['check sign'] = ''
    control['result'] = ''

    val_year_idx = control[control.iloc[:, 0] == 'Val Year'].index
    if not val_year_idx.empty:
        idx = val_year_idx[0]
        control.at[idx, 'check sign'] = 'Check Sign'
        control.at[idx, 'result'] = check_sign_total

    columns_name_argo = list(cf_argo.columns[:2])
    columns_cf_argo =  columns_name_argo + cols_to_compare
    columns_cf_argo = [k for k in columns_cf_argo if k in cf_argo.columns]
    cf_argo = cf_argo[columns_cf_argo]

    columns_name_rafm = list(cf_rafm.columns[:2])
    columns_cf_rafm =  columns_name_rafm + cols_to_compare
    columns_cf_rafm = [k for k in columns_cf_rafm if k in cf_rafm.columns]
    cf_rafm = cf_rafm[columns_cf_rafm]

    return {
        'Control': control,
        'Code': code,
        "CF ARGO REAS": cf_argo,
        "RAFM Output REAS": cf_rafm,
        "Checking Summary REAS": final
    }

if __name__ == '__main__':
    import multiprocessing
    multiprocessing.freeze_support()