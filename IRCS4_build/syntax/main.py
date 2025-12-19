import os
import pandas as pd
from xlsxwriter.utility import xl_col_to_name
import syntax.control_4_trad as trad
import syntax.control_4_ul as ul
import syntax.control_4_reas as reas
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings
import shutil
import datetime

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

cols_to_sum_dict = {
    'trad': trad.cols_to_compare,
    'ul': ul.columns_to_sum_argo,
    'reas': reas.cols_to_compare
}

def auto_adjust_column_width_xlwings(ws, df_sheet, sample_size=100):
    if not hasattr(df_sheet, 'columns'):
        return
    header_offset = 0
    if isinstance(df_sheet, pd.DataFrame) and df_sheet.shape[1] > 0:
        header_offset = 1
    for i, col in enumerate(df_sheet.columns):
        try:
            samples = []
            samples.append(str(col))
            sample_vals = df_sheet[col].head(sample_size).astype(str).tolist()
            samples.extend(sample_vals)
            max_len = max(len(s) for s in samples)
            adjusted_width = max(8, max_len + 2)
            try:
                ws.range((1, i + 1)).column_width = adjusted_width
            except Exception:
                try:
                    ws.api.Columns(i + 1).ColumnWidth = adjusted_width
                except Exception:
                    pass
        except Exception:
            try:
                ws.range((1, i + 1)).column_width = 12
            except Exception:
                pass


def apply_number_formats_xlwings(ws, df_sheet):
    if not hasattr(df_sheet, 'columns'):
        return
    nrows = len(df_sheet) + 1
    ncols = len(df_sheet.columns)
    for col_idx, col_name in enumerate(df_sheet.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_name_lower = str(col_name).lower()
        if 'speed duration' in col_name_lower:
            number_format = '@'
        elif 'include year' in col_name_lower or 'exclude year' in col_name_lower:
            number_format = '0'
        else:
            number_format = '_-* #,##0_-;_-* (#,##0);_-* "-"_-;_-@_-'
        try:
            rng = ws.range(f"{col_letter}2:{col_letter}{nrows}")
            rng.number_format = number_format
        except Exception:
            try:
                ws.api.Range(f"{col_letter}2:{col_letter}{nrows}").NumberFormat = number_format
            except Exception:
                pass


def apply_border_xlwings(ws, df_sheet):
    try:
        nrows = len(df_sheet) + 1
        ncols = len(df_sheet.columns)
        if nrows < 1 or ncols < 1:
            return
        last_cell = f"{get_column_letter(ncols)}{nrows}"
        full_range = ws.range(f"A1:{last_cell}")
        full_range.api.Borders.LineStyle = 1
        full_range.api.Borders.Weight = 2
    except Exception:
        try:
            for r in range(1, nrows + 1):
                for c in range(1, ncols + 1):
                    try:
                        cell = ws.api.Cells(r, c)
                        cell.Borders.LineStyle = 1
                        cell.Borders.Weight = 2
                    except Exception:
                        pass
        except Exception:
            pass


def apply_accounting_to_all_xlwings(ws, df_sheet, start_row=2):
    nrows = len(df_sheet) + start_row - 1
    ncols = len(df_sheet.columns)

    for col_idx in range(1, ncols+1):
        try:
            for row_idx in range(start_row, nrows+1):
                cell = ws.api.Cells(row_idx, col_idx)
                cell.NumberFormat = '_-* #,##0_-;_-* (#,##0);_-* "-"_-;_-@_-'
        except Exception as e:
            print(f"⚠️ Failed to apply accounting format to cell ({row_idx}, {col_idx}): {e}")


def write_checking_summary_formulas_xlwings(ws, df_sheet, jenis, start_row=2):
    sheet_names = {
        'trad': {
            'cf_argo': 'CF ARGO AZTRAD',
            'cf_rafm': 'RAFM Output AZTRAD',
            'rafm_manual': 'RAFM Output Manual',
            'uvsg': 'RAFM Output AZUL_PI'
        },
        'ul': {
            'cf_argo': 'CF ARGO AZUL',
            'cf_rafm': 'RAFM Output AZUL',
            'rafm_manual': 'RAFM Output Manual'
        },
        'reas': {
            'cf_argo': 'CF ARGO REAS',
            'cf_rafm': 'RAFM Output REAS',
            'rafm_manual': 'RAFM Output Manual'
        }
    }

    nrows = len(df_sheet)
    ncols = len(df_sheet.columns)
    if jenis == 'trad':
        start_col_idx = 5
        cf_argo_col_offset = 3
        cf_rafm_col_offset = 7
        rafm_manual_col_offset = 7
        uvsg_col_offset = 7
    elif jenis == 'ul':
        start_col_idx = 4
        cf_argo_col_offset = 3
        cf_rafm_col_offset = 6
        rafm_manual_col_offset = 6
    else:
        start_col_idx = 4
        cf_argo_col_offset = 3
        cf_rafm_col_offset = 3
        rafm_manual_col_offset = 3

    for row_idx in range(nrows):
        row_excel = start_row + row_idx
        for col_idx in range(start_col_idx, ncols + 1):
            relative_offset = col_idx - start_col_idx
            if jenis == 'trad':
                cf_argo_col = get_column_letter(cf_argo_col_offset + relative_offset)
                cf_rafm_col = get_column_letter(cf_rafm_col_offset + relative_offset)
                rafm_manual_col = get_column_letter(rafm_manual_col_offset + relative_offset)
                uvsg_col = get_column_letter(uvsg_col_offset + relative_offset)
                formula = (
                    f"='{sheet_names['trad']['cf_argo']}'!{cf_argo_col}{row_excel}"
                    f"-'{sheet_names['trad']['cf_rafm']}'!{cf_rafm_col}{row_excel}"
                    f"+'{sheet_names['trad']['rafm_manual']}'!{rafm_manual_col}{row_excel}"
                    f"-'{sheet_names['trad']['uvsg']}'!{uvsg_col}{row_excel}"
                )
            elif jenis == 'ul':
                cf_argo_col = get_column_letter(cf_argo_col_offset + relative_offset)
                cf_rafm_col = get_column_letter(cf_rafm_col_offset + relative_offset)
                rafm_manual_col = get_column_letter(rafm_manual_col_offset + relative_offset)
                formula = (
                    f"='{sheet_names['ul']['cf_argo']}'!{cf_argo_col}{row_excel}"
                    f"-'{sheet_names['ul']['cf_rafm']}'!{cf_rafm_col}{row_excel}"
                    f"-'{sheet_names['ul']['rafm_manual']}'!{rafm_manual_col}{row_excel}"
                )
            else:
                cf_argo_col = get_column_letter(cf_argo_col_offset + relative_offset)
                cf_rafm_col = get_column_letter(cf_rafm_col_offset + relative_offset)
                rafm_manual_col = get_column_letter(rafm_manual_col_offset + relative_offset)
                formula = (
                    f"='{sheet_names['reas']['cf_argo']}'!{cf_argo_col}{row_excel}"
                    f"-'{sheet_names['reas']['cf_rafm']}'!{cf_rafm_col}{row_excel}"
                    f"+'{sheet_names['reas']['rafm_manual']}'!{rafm_manual_col}{row_excel}"
                )
            ws.range(f"{get_column_letter(col_idx)}{row_excel}").formula = formula


def add_sheets_to_rafm_manual(rafm_manual_path, result_dict, output_path, output_filename, jenis):
    app = None
    wb = None
    try:
        if not os.path.exists(rafm_manual_path):
            print(f"❌ RAFM Manual not found: {rafm_manual_path}")
            return None

        print(f"\n🚀 Create Excel File...")
        start_time = time.time()
        os.makedirs(output_path, exist_ok=True)
        output_file = os.path.join(output_path, output_filename)

        if os.path.exists(output_file):
            try:
                os.remove(output_file)
            except PermissionError:
                try:
                    os.remove(output_file)
                except PermissionError:
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    base_name = os.path.splitext(output_filename)[0]
                    ext = os.path.splitext(output_filename)[1]
                    output_filename = f"{base_name}_{timestamp}{ext}"
                    output_file = os.path.join(output_path, output_filename)

        shutil.copy2(rafm_manual_path, output_file)
        time.sleep(0.5)

        app = xw.App(visible=True)
        app.display_alerts = False
        app.screen_updating = False
        wb = app.books.open(output_file)

        sheet_names_list = [sh.name for sh in wb.sheets]
        if 'Sheet1' in sheet_names_list and 'RAFM Output Manual' not in sheet_names_list:
            wb.sheets['Sheet1'].name = 'RAFM Output Manual'
        elif 'Sheet1' in sheet_names_list and 'RAFM Output Manual' in sheet_names_list:
            wb.sheets['Sheet1'].delete()

        if jenis == 'trad':
            sheet_order = [
                'Control', 'Code',
                'CF ARGO AZTRAD', 'RAFM Output AZTRAD',
                'RAFM Output Manual',
                'RAFM Output AZUL_PI',
                'Checking Summary AZTRAD'
            ]
        elif jenis == 'ul':
            sheet_order = [
                'Control', 'Code',
                'CF ARGO AZUL', 'RAFM Output AZUL',
                'RAFM Output Manual',
                'Checking Summary AZUL'
            ]
        else:
            sheet_order = [
                'Control', 'Code',
                'CF ARGO REAS', 'RAFM Output REAS',
                'RAFM Output Manual',
                'Checking Summary REAS'
            ]

        for sheet_name, df in result_dict.items():
            if sheet_name == 'RAFM Output Manual':
                continue

            df = df.copy().replace({pd.NA: None, pd.NaT: None}).where(pd.notna(df), None)

            if sheet_name in [sh.name for sh in wb.sheets]:
                wb.sheets[sheet_name].delete()

            ws = wb.sheets.add(sheet_name, after=wb.sheets[-1])

            if sheet_name == 'Control':
                ws.range('A1').value = df.values.tolist()
            else:
                data_with_header = [df.columns.tolist()] + df.values.tolist()
                ws.range('A1').value = data_with_header

                apply_number_formats_xlwings(ws, df)
                apply_border_xlwings(ws, df)
                auto_adjust_column_width_xlwings(ws, df)

                if sheet_name.startswith("Checking Summary"):
                    write_checking_summary_formulas_xlwings(ws, df, jenis, start_row = 2)
                    apply_accounting_to_all_xlwings(ws, df, start_row = 2)

            try:
                ws.autofit(axis='columns')
            except Exception:
                pass

        for target_idx, sheet_name in enumerate(sheet_order):
            if sheet_name in [sh.name for sh in wb.sheets]:
                current_idx = [sh.name for sh in wb.sheets].index(sheet_name)
                if current_idx != target_idx:
                    wb.sheets[sheet_name].api.Move(Before=wb.sheets[target_idx].api)
        try:
            wb.api.RefreshAll()
            wb.app.api.CalculateFullRebuild()
            wb.app.api.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass

        wb.save()
        time.sleep(1.5)

        wb.close()
        app.quit()
        app = None
        wb = None

        print(f"✅ Done : {output_file}")
        return output_file

    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback; traceback.print_exc()
        return None

    finally:
        if wb is not None:
            try:
                wb.close()
            except:
                pass
        if app is not None:
            try:
                app.quit()
            except:
                pass
        time.sleep(1)


def process_input_file_threadsafe(file_path):
    filename = os.path.basename(file_path).lower()
    if 'trad' in filename:
        jenis = 'trad'
        result = trad.main({"input excel": file_path})
    elif 'ul' in filename:
        jenis = 'ul'
        result = ul.main({"input excel": file_path})
    elif 'reas' in filename:
        jenis = 'reas'
        result = reas.main({"input excel": file_path})
    else:
        return (file_path, None, None)
    return (file_path, jenis, result)


def main(input_path):
    print("\n" + "="*60)
    print("🔧 CONTROL 4")
    print("="*60)
    start_time = time.time()
    if os.path.isfile(input_path):
        files = [input_path]
    elif os.path.isdir(input_path):
        files = [os.path.join(input_path, f) for f in os.listdir(input_path)
                 if f.endswith(".xlsx") and not f.startswith("~$")]
    else:
        print(f"❌ Path not valid: {input_path}")
        return
    if not files:
        print("📂 There is no file .xlsx")
        return
    print(f"📊 Memproses {len(files)} file\n")
    compute_results = []
    with ThreadPoolExecutor(max_workers=min(4, len(files))) as executor:
        future_to_file = {executor.submit(process_input_file_threadsafe, f): f for f in files}
        for future in as_completed(future_to_file):
            try:
                file_path, jenis, result = future.result()
                if jenis and result:
                    compute_results.append((file_path, jenis, result))
            except Exception as e:
                print(f"❌ Error processing {future_to_file[future]}: {e}")
    success_count = 0
    fail_count = 0
    for file_path, jenis, result in compute_results:
        try:
            filename = os.path.basename(file_path)
            print(f"\n📄 Process to Excel: {filename} ({jenis})")
            df = pd.read_excel(file_path, sheet_name='File Path')
            df.columns = df.columns.str.strip()
            df['Name'] = df['Name'].astype(str).str.strip().str.lower()
            df['File Path'] = df['File Path'].astype(str).str.strip()
            output_path = df.loc[df['Name'] == 'output_path', 'File Path'].values[0]
            output_filename = df.loc[df['Name'] == 'output_filename', 'File Path'].values[0]
            rafm_manual_path = df.loc[df['Name'] == 'rafm manual', 'File Path'].values[0]
            output_file = add_sheets_to_rafm_manual(
                rafm_manual_path, result, output_path, output_filename, jenis
            )
            if output_file:
                success_count += 1
            else:
                fail_count += 1
        except Exception as e:
            print(f"❌ Error Excel stage: {e}")
            import traceback; traceback.print_exc()
            fail_count += 1
    elapsed = time.time() - start_time
    print("\n" + "="*60)
    print(f"⏱️ Total Runtime: {elapsed:.2f} detik")
    print(f"📊 Total: {len(files)} file(s)")
    print(f"✅ Success: {success_count}")
    print(f"❌ Failed: {fail_count}")
    if len(files) > 0:
        print(f"⚡ Avg: {elapsed/len(files):.2f} Second/file")
    print("="*60)


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        main(sys.argv[1])
    else:
        print("Usage: python main.py <input_file_or_folder>")
